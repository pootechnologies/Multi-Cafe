
from rest_framework.views import APIView
from django.db.models import Sum, Count
from django.http import HttpResponse
from rest_framework.permissions import BasePermission
from rest_framework.parsers import MultiPartParser
from django.shortcuts import get_object_or_404
from django.db.models import F, Sum, ExpressionWrapper, DecimalField
from django.db import IntegrityError
from django.utils import timezone
from datetime import timedelta
from django.utils.dateparse import parse_date
import calendar
import openpyxl

from rest_framework import generics, status, permissions, filters
from rest_framework.response import Response

from tenants.views import IsTenantOwnerOrAdmin, IsTenantUser
from .models import (
    Product, Supplier, Order, OrderItem, Category, 
    CustomerInfo, CompanyInfo, OrderLog, Report, ExpenseTypes, 
    OtherExpenses, OrderPaymentLog, ProductLog
)
from .serializers import (
    CategorySerializer, 
                        ProductSerializer,
                        ProductGetSerializer, 
                        SupplierSerializer,
                        ProductGetReportSerializer,
                        OrderSerializer, 
                        OrderLightSerializer,
                        OrderItemSerializer, 
                        CategorySerializer, 
                        CustomerInfoSerializer,
                        CompanyInfoSerializer,
                        OrderLogSerializer,
                        OrderReportSerializer,
                        ExpenseTypesSerializer,
                        OtherExpensesSerializer,
                        OtherExpensesGetSerializer,
                        OrderPaymentLogSerializer,
                        ProductLogSerializer,
                          )
from rest_framework.pagination import PageNumberPagination
from django.db.models import Q
from django.core.exceptions import ValidationError
from .utils import create_order_log
class Pagination(PageNumberPagination):
    page_size = 10  # default items per page
    page_size_query_param = 'page_size'  # allow client to override
    max_page_size = 100
class SupplierListCreateView(generics.ListCreateAPIView):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]

    def get(self, request, format=None):
        try:
            
            supplier = Supplier.objects.all()
            serializer = SupplierSerializer(supplier, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)                            
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Supplier.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        supplier = serializer.save()
        headers = self.get_success_headers(serializer.data)
        return Response(
            {
                "message": "Supplier created successfully",
                "supplier": SupplierSerializer(supplier).data
            },
            status=status.HTTP_201_CREATED,
            headers=headers
        )
    
class SupplierRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        supplier = serializer.save()
        return Response(
            {
                "message": "Supplier updated successfully",
                "supplier": SupplierSerializer(supplier).data
            },
            status=status.HTTP_200_OK
        )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(
            {
                "message": "Supplier deleted successfully"
            },
            status=status.HTTP_204_NO_CONTENT
        )    
class CategoryListCreateView(generics.ListCreateAPIView):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]

    def get(self, request, format=None):
        try:
            
            category = Category.objects.all().order_by('id')
            serializer = CategorySerializer(category, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK) 
       
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Category.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        category = serializer.save()
        headers = self.get_success_headers(serializer.data)
        return Response(
            {
                "message": "Category created successfully",
                "category": CategorySerializer(category).data
            },
            status=status.HTTP_201_CREATED,
            headers=headers
        )
class CategoryRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    # 
    # lookup_field = 'id'
    # 
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        category = serializer.save()
        return Response(
            {
                "message": "Category updated successfully",
                "category": CategorySerializer(category).data
            },
            status=status.HTTP_200_OK
        )
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(
            {
                "message": "Category deleted successfully"
            },
            status=status.HTTP_204_NO_CONTENT
        )
    

class ProductListCreateView(generics.ListCreateAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request, format=None):
        try:
            

            search_query = request.query_params.get('search', None)
            include_all = request.query_params.get('include_all', '').lower() in ('1', 'true', 'yes')

            products = Product.objects.all()
            
            # Apply search
            if search_query:
                products = products.filter(
                    Q(name__icontains=search_query) |
                    Q(category__name__icontains=search_query) |
                    Q(specification__icontains=search_query)
                )

            # Ensure consistent ordering for pagination
            products = products.order_by('-id')  # or '-id'

            # Paginate
            paginator = Pagination()
            page = paginator.paginate_queryset(products, request)
            if page is not None:
                page_data = ProductGetSerializer(page, many=True).data
                if include_all:
                    all_data = ProductGetSerializer(products, many=True).data
                    return Response({
                        'count': paginator.page.paginator.count,
                        'next': paginator.get_next_link(),
                        'previous': paginator.get_previous_link(),
                        'results': page_data,
                        'all_results': all_data,
                    })
                return paginator.get_paginated_response(page_data)

            # Fallback - no pagination applied
            serializer = ProductGetSerializer(products, many=True)
            return Response(serializer.data)
              
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Product.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        product = serializer.save()
        headers = self.get_success_headers(serializer.data)
        return Response(
            {
                "message": "Product created successfully",
                "product": ProductSerializer(product).data
            },
            status=status.HTTP_201_CREATED,
            headers=headers
        )
class ProductRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Product.objects.order_by('id')
    serializer_class = ProductSerializer
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        product = serializer.save()
        return Response(
            {
                "message": "Product updated successfully",
                "product": ProductSerializer(product).data
            },
            status=status.HTTP_200_OK
        )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(
            {
                "message": "Product deleted successfully"
            },
            status=status.HTTP_204_NO_CONTENT
        )    

# //////////////////////////////////////////////////////////////////
class CustomerListCreateAPIView(APIView):
    # permission_classes = (permissions.AllowAny,)
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request, format=None):
        try:

            search_query = request.query_params.get('search', None)
            include_all = request.query_params.get('include_all', '').lower() in ('1', 'true', 'yes')

            customers = CustomerInfo.objects.all()
            
            # Apply search
            if search_query:
                customers = customers.filter(
                    Q(name__icontains=search_query)
                )

            # Ensure consistent ordering for pagination
            customers = customers.order_by('-id')  # or '-id'

            # Paginate
            paginator = Pagination()
            page = paginator.paginate_queryset(customers, request)
            if page is not None:
                page_data = CustomerInfoSerializer(page, many=True).data
                if include_all:
                    all_data = CustomerInfoSerializer(customers, many=True).data
                    return Response({
                        'count': paginator.page.paginator.count,
                        'next': paginator.get_next_link(),
                        'previous': paginator.get_previous_link(),
                        'results': page_data,
                        'all_results': all_data,
                    })
                return paginator.get_paginated_response(page_data)

            # Fallback - no pagination applied
            serializer = CustomerInfoSerializer(customers, many=True)
            return Response(serializer.data)
                  
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Customers.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def post(self, request, format=None):
        try:
                        
            serializer = CustomerInfoSerializer(data=request.data)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            validated_data = serializer.validated_data
            # serializer.create(validated_data, user=request.user)
            serializer.create(validated_data)
            return Response({"message": f"Customer Created successfully."}, status=status.HTTP_201_CREATED)     
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while creating the Customer.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
class CustomerRetrieveUpdateDeleteAPIView(APIView):
    # permission_classes = (permissions.AllowAny,)
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request, pk):
        try:
                          
            if not CustomerInfo.objects.filter(id=pk).exists():
                return Response(
                    {"error": "Customer Does not Exist."},
                    status=status.HTTP_404_NOT_FOUND
                )
            customer = CustomerInfo.objects.get(id=pk)
            serializer = CustomerInfoSerializer(customer)
            return Response(serializer.data, status=status.HTTP_200_OK)      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Customer.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def put(self, request, pk):
        try:
                          
            if not CustomerInfo.objects.filter(id=pk).exists():
                return Response(
                    {"error": "Customer Does not Exist."},
                    status=status.HTTP_404_NOT_FOUND
                )
            customer = CustomerInfo.objects.get(id=pk)
            serializer = CustomerInfoSerializer(data=request.data)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            validated_data = serializer.validated_data
            serializer.update(customer, validated_data)
            return Response({"message": f"Customer Updated successfully."}, status=status.HTTP_200_OK)      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while updating the Customer.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
       
    def patch(self, request, pk):
        try:
                           
            if not CustomerInfo.objects.filter(id=pk).exists():
                return Response(
                    {"error": "Customer Does not Exist."},
                    status=status.HTTP_404_NOT_FOUND
                )
            customer = CustomerInfo.objects.get(id=pk) 
            serializer = CustomerInfoSerializer (customer, data=request.data, partial=True)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            serializer.save()
            return Response({"message": f"Customer Updated successfully."}, status=status.HTTP_200_OK)      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while updating the Customer.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def delete(self, request, pk):
        try:
                         
            if not CustomerInfo.objects.filter(id=pk).exists():
                return Response(
                    {"error": "Customer Does not Exist."},
                    status=status.HTTP_404_NOT_FOUND
                )
            CustomerInfo.objects.get(id=pk).delete()
            if not CustomerInfo.objects.filter(id=pk).exists():
                return Response({"message": f"Customer Deleted successfully."},
                    status=status.HTTP_204_NO_CONTENT
                )
            else:
                return Response(
                    {"error": "Failed to delete an Customer."},
                    status=status.HTTP_400_BAD_REQUEST
                )      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Deleting the Customer.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CompanyListCreateAPIView(APIView):
    # permission_classes = (permissions.AllowAny,)
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request, format=None):
        try:
            
            company = CompanyInfo.objects.all()
            serializer = CompanyInfoSerializer(company, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
            
                      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Company.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def post(self, request, format=None):
        try:
             
            serializer = CompanyInfoSerializer(data=request.data)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            validated_data = serializer.validated_data
            validated_data['user'] = user
            serializer.create(validated_data, user=request.user)
            return Response({"message": f"Company Created successfully."}, status=status.HTTP_201_CREATED)

        except KeyError as e:
            print(e)
            return Response(
                {"error": f"An error occurred while Creating the Company.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class CompanyRetrieveUpdateDeleteAPIView(APIView):
    # permission_classes = (permissions.AllowAny,)
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request, pk):
        try:
                         
            if not CompanyInfo.objects.filter(id=pk).exists():
                return Response(
                    {"error": "Company Does not Exist."},
                    status=status.HTTP_404_NOT_FOUND
                )
            company = CompanyInfo.objects.get(id=pk)
            serializer = CompanyInfoSerializer(company)
            return Response(serializer.data, status=status.HTTP_200_OK)      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Company.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def put(self, request, pk):
        try:
           
            if not CompanyInfo.objects.filter(id=pk).exists():
                return Response(
                    {"error": "Company Does not Exist."},
                    status=status.HTTP_404_NOT_FOUND
                )
            company = CompanyInfo.objects.get(id=pk)
            serializer = CompanyInfoSerializer(data=request.data)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            validated_data = serializer.validated_data
            serializer.update(company, validated_data)
            return Response({"message": f"Company Updated successfully."}, status=status.HTTP_200_OK)        
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while updating the Company.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def patch(self, request, pk):
        try:
                     
            if not CompanyInfo.objects.filter(id=pk).exists():
                return Response(
                    {"error": "Company Does not Exist."},
                    status=status.HTTP_404_NOT_FOUND
                )
            company = CompanyInfo.objects.get(id=pk)    
            serializer = CompanyInfoSerializer(company, data=request.data, partial=True)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            serializer.save()
            return Response({"message": f"Company Updated successfully."}, status=status.HTTP_200_OK)      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while updating the Company.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def delete(self, request, pk):
        try:
                          
            if not CompanyInfo.objects.filter(id=pk).exists():
                return Response(
                    {"error": "Company Does not Exist."},
                    status=status.HTTP_404_NOT_FOUND
                )
            CompanyInfo.objects.get(id=pk).delete()
            if not CompanyInfo.objects.filter(id=pk).exists():
                return Response({"message": f"Company Deleted successfully."},
                    status=status.HTTP_204_NO_CONTENT
                )
            else:
                return Response(
                    {"error": "Failed to delete an Company."},
                    status=status.HTTP_400_BAD_REQUEST
                )      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Deleting the Company.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class OrderPermission(BasePermission):
    def has_permission(self, request, view):
        user = request.user
        return user and (getattr(user, "role", None) == "Manager" or user.is_superuser or user.role == 'Salesman' or user.role == 'Sales Manager')


class OrderListCreatView(generics.ListCreateAPIView):
    queryset = Order.objects.filter(credit=False).order_by('-id')
    # permission_classes = [OrderPermission]
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    serializer_class = OrderSerializer
    pagination_class = Pagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['=customer__name', '=payment_status']  # 🔍 allow searching by customer's name and payment status

    def get_queryset(self):
        # only fetch id, name, email from the DB
        return Order.objects.only('id', 'customer', 'customer__name', 'status', 'receipt', 'receipt_id', 'order_date', 'sub_total', 'vat',  'total_amount', 'payment_status', 'paid_amount', 'unpaid_amount', 'credit', 'user').select_related('customer').filter(credit=False).order_by('-id')


    def get_serializer_class(self):
        # Use light serializer for list, full serializer for create
        if self.request.method == 'GET':
            return OrderLightSerializer
        return OrderSerializer
    
    def list(self, request, *args, **kwargs):
        # Get the queryset
        queryset = self.filter_queryset(self.get_queryset())
        
        # Get paginated data
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            data = self.get_paginated_response(serializer.data).data
        else:
            data = self.get_serializer(queryset, many=True).data
            data = {
                'count': len(data),
                'results': data
            }
        
        # Add all results
        all_serializer = self.get_serializer(queryset, many=True)
        data['all_results'] = all_serializer.data
        
        return Response(data)

    # def perform_create(self, serializer):
    #     serializer.save(user=self.request.user)

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        id = response.data.get('receipt_id')
        return Response({
            "message": "Order created successfully.",
            "data": response.data,
            "id": id
        }, status=status.HTTP_201_CREATED)


class OrderDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]

    def get(self, request, *args, **kwargs):
        response = super().get(request, *args, **kwargs)
        id = response.data.get('receipt_id')
        return Response({
            "message": "Order retrived successfully.",
            "data": response.data,
            "id": id
        }, status=status.HTTP_200_OK)

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        return Response({
            "message": "Order updated successfully.",
            "data": response.data
        }, status=status.HTTP_200_OK)
    
    def destroy(self, request, *args, **kwargs):
        super().destroy(request, *args, **kwargs)
        return Response({"message": "Order Deleted successfully."}, status=status.HTTP_200_OK)

class OrderItemListCreateView(generics.ListCreateAPIView):
    queryset = OrderItem.objects.filter(order__credit=False).order_by('id')
    serializer_class = OrderItemSerializer
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        return Response({
            "message": "Order Item created successfully.",
            "data": response.data
        }, status=status.HTTP_201_CREATED)

class OrderItemDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = OrderItem.objects.all()
    serializer_class = OrderItemSerializer
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        return Response({
            "message": "Order Item updated successfully.",
            "data": response.data
        }, status=status.HTTP_200_OK)
    

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        order = instance.order  # Get the order before deletion
        self.perform_destroy(instance)  # Delete the OrderItem

        # Check if the order still exists
        if Order.objects.filter(pk=order.pk).exists():
            items_count = OrderItem.objects.filter(order=order).count()
            return Response({
                "message": "Order Item Deleted successfully.",
                "order_id": order.id,
                "number_of_items": items_count
            }, status=status.HTTP_200_OK)
        else:
            return Response({
                "message": "Order Item Deleted successfully. The order was also deleted because it had no items left.",
                "order_id": None,
                "number_of_items": 0
            }, status=status.HTTP_200_OK)


class OrderCreditListAPIView(generics.ListAPIView):
    queryset = Order.objects.filter(credit=True).order_by('-id')
    # permission_classes = [OrderPermission]
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    serializer_class = OrderLightSerializer
    pagination_class = Pagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['=customer__name', '=payment_status']  # 🔍 allow searching by customer's name and payment status

    def get_queryset(self):
        # only fetch id, name, email from the DB
        return Order.objects.only('id', 'customer', 'customer__name', 'status', 'receipt', 'receipt_id', 'order_date', 'sub_total', 'vat',  'total_amount', 'payment_status', 'paid_amount', 'unpaid_amount', 'credit', 'user').select_related('customer').filter(credit=True).order_by('-id')



class OrderItemCreditListView(generics.ListAPIView):
    queryset = OrderItem.objects.filter(order__credit=True).order_by('id')
    serializer_class = OrderItemSerializer
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    # def get_queryset(self):
    #     # only fetch id, name, email from the DB
    #     return OrderItem.objects.only('id', 'order', 'product', 'product_price', 'product__name', 'item_receipt', 'package', 'unit', 'quantity', 'unit_price', 'price', 'status').select_related('order', 'product').filter(order__credit=True).order_by('id')

class RetriveRevenueAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request): 
        try:
           
     
            revenue = Order.objects.filter(status="Done", payment_status='Paid').aggregate(total_revenue=Sum('total_amount'))        
            return Response(revenue, status=status.HTTP_200_OK)         
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Revenue.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class RetriveSalesPersonRevenueAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request): 
        try:
            
            orders = Order.objects.filter(user_email=user.email, status='Done', payment_status='Paid')
            revenue = orders.aggregate(total_revenue=Sum('total_amount'))

            return Response(revenue, status=status.HTTP_200_OK)   
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Revenue.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class RetriveTotalOrdersAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request): 
        try:
           
            orders = Order.objects.filter(user_email=user.email, status='Done', payment_status='Paid')
            total_orders = orders.aggregate(total_orders=Count('total_amount'))

            return Response(total_orders, status=status.HTTP_200_OK)   
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Revenue.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class RetriveProfitAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request): 
        try:
            
            # Get all paid and done orders
            paid_done_orders = Order.objects.filter(status="Done", payment_status='Paid')
            
            # Calculate revenue from these orders
            revenue = paid_done_orders.aggregate(total_revenue=Sum('total_amount'))
            print(revenue)
            
            # Get all order items from these orders
            order_items = OrderItem.objects.filter(order__in=paid_done_orders)

            # Calculate cost and profit from these order items
            cost = order_items.aggregate(total_cost=Sum('cost'))
            print(cost)
            
            # profit = order_items.aggregate(
            #     total_profit=Sum(revenue['total_revenue'] - cost['total_cost'])
            # )
            # profit = profit or 0
            # if profit['total_profit'] is None:
            #     profit['total_profit'] = 0

            if revenue['total_revenue'] is None or cost['total_cost'] is None:
                profit = {'total_profit': 0.00}
            else:
                profit = {'total_profit': float(revenue['total_revenue']) - float(cost['total_cost'])}

            print(profit)
            return Response(profit, status=status.HTTP_200_OK)        
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Profit. {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class OrderReceiptAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request, pk):
        try:
            # Retrieve the order along with related data
            order = get_object_or_404(Order.objects.prefetch_related('items__product').select_related('customer'), id=pk)
            
            # Get the first (or default) company info (assuming there's only one)
            company = CompanyInfo.objects.first()

            # Prepare receipt data
            receipt_data = {
                "company": {
                    "en_name": company.en_name if company else "Company",
                    "am_name": company.am_name if company else "Company",
                    "email": company.email if company else "",
                    "phone": company.phone1 if company else "",
                    "tin_number": company.tin_number if company else "",
                    "vat_number": company.vat_number if company else "",
                    "bank_accounts": company.bank_accounts if company else {},
                    "country": company.country if company else "",
                    "region": company.region if company else "",
                    "zone": company.zone if company else "",
                    "city": company.city if company else "",
                    "sub_city": company.sub_city if company else "",
                    "logo": company.logo.url if company and company.logo else None
                },
                "customer": {
                    "name": order.customer.name if order.customer else "Customer",
                    "phone": order.customer.phone if order.customer else "",
                    "tin_number": order.customer.tin_number if order.customer else "",
                    "vat_number": order.customer.vat_number if order.customer else "",
                    "customer_fs": order.customer.fs_number if order.customer else "",
                    "zone": order.customer.zone if order.customer else "",
                    "city": order.customer.city if order.customer else "",
                    "sub_city": order.customer.sub_city if order.customer else "",
                },
                "order_details": {
                    "order_id": order.id,
                    "date": order.order_date,
                    "status": order.status,
                    "receipt": order.receipt,
                    "sub_total": order.sub_total,
                    "vat": order.vat,
                    "total_amount": order.total_amount,
                    "user": order.user if order.user else "User",
                },
                "items": [
                    {
                        "product_name": item.product.name if item.product else "Unknown",
                        "specification": item.product.specification if item.product else "No Spec",
                        "quantity": item.quantity,
                        "product_price": item.unit_price if item.unit_price else item.product.selling_price if item.product else 0,
                        "unit": item.unit,
                        "unit_price": item.unit_price,
                        "price": item.price,
                        "status": item.status,
                    }
                    for item in order.items.all()
                ]
            }

            return Response(receipt_data, status=status.HTTP_200_OK)
        except Order.DoesNotExist:
            return Response(
                {"error": "Order does not exist."},
                status=status.HTTP_404_NOT_FOUND
            )
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Order Receipt.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class OrderLogAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request):
        try:
           
            
            search_query = request.query_params.get('search', None)
            include_all = request.query_params.get('include_all', '').lower() in ('1', 'true', 'yes')

            order_log = OrderLog.objects.all()
            
            # Apply search
            if search_query:
                order_log = order_log.filter(
                    Q(object_id__icontains=search_query)
                )

            # Ensure consistent ordering for pagination
            order_log = order_log.order_by('-id')  # or '-id'

            # Paginate
            paginator = Pagination()
            page = paginator.paginate_queryset(order_log, request)
            if page is not None:
                page_data = OrderLogSerializer(page, many=True).data
                if include_all:
                    all_data = OrderLogSerializer(order_log, many=True).data
                    return Response({
                        'count': paginator.page.paginator.count,
                        'next': paginator.get_next_link(),
                        'previous': paginator.get_previous_link(),
                        'results': page_data,
                        'all_results': all_data,
                    })
                return paginator.get_paginated_response(page_data)

            # Fallback - no pagination applied
            serializer = OrderLogSerializer(order_log, many=True)
            return Response(serializer.data)

        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Order Log.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



class ExcelReportAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request):
        try:
          

            # Get query parameters
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')

            # Start with base queryset
            reports = Report.objects.all()

            # Apply date range filter if both dates are provided
            if start_date and end_date:
                # Convert string dates to datetime objects for comparison
                start = parse_date(start_date)
                end = parse_date(end_date)

                if start and end:
                    # Include records up to the end of the day on end_date
                    reports = reports.filter(
                        order_date__gte=start,
                        order_date__lt=end + timedelta(days=1)
                    )

            serializer = OrderReportSerializer(reports, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"An error occurred while retrieving the Order Report: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



class ListOutOFStockProductAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request):
        try:
            
            out_of_stock_products = Product.objects.filter(stock__lte=3)
            serializer = ProductGetSerializer(out_of_stock_products, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Stock Shortage.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class CountNearExpirationDateProductAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request):
        try:
            
            out_of_stock_products = Product.objects.filter(stock__lte=3).aggregate(out_of_stock=Count('name'))
            return Response(out_of_stock_products, status=status.HTTP_200_OK)

        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Stock Shortage.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ExpenseTypesListCreateAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]

    # permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        try:
            
            expense_type = ExpenseTypes.objects.all().order_by('id')
            serializer = ExpenseTypesSerializer(expense_type, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)              
                      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Expense Types.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def post(self, request, format=None):
        try:
            

            serializer = ExpenseTypesSerializer(data=request.data)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            validated_data = serializer.validated_data
            # serializer.create(validated_data, user=request.user)
            serializer.create(validated_data)
            return Response({"message": f"Expense Types created successfully."}, status=status.HTTP_201_CREATED)
                      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while creating the Expense Types.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ExpenseTypesRetrieveUpdateDeleteAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    # permission_classes = (permissions.AllowAny,)
    def get(self, request, pk):
        try:
           
            if not ExpenseTypes.objects.filter(id=pk).exists():
                return Response(
                    {"error": "Expense Types Does not Exist."},
                    status=status.HTTP_404_NOT_FOUND
                )
            expense_type = ExpenseTypes.objects.get(id=pk)
            serializer = ExpenseTypesSerializer(expense_type)
            return Response(serializer.data, status=status.HTTP_200_OK)     
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Expense Types.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def put(self, request, pk):
        try:
                        
            if not ExpenseTypes.objects.filter(id=pk).exists():
                return Response(
                    {"error": "Expense Types Does not Exist."},
                    status=status.HTTP_404_NOT_FOUND
                )
            expense_type = ExpenseTypes.objects.get(id=pk)
            serializer = ExpenseTypesSerializer(data=request.data)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            validated_data = serializer.validated_data
            serializer.update(expense_type, validated_data)
            return Response({"message": f"Expense Types Updated successfully."}, status=status.HTTP_200_OK)      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while updating the Expense Types.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
    def patch(self, request, pk):
        try:
                         
            if not ExpenseTypes.objects.filter(id=pk).exists():
                return Response(
                    {"error": "Expense Types Does not Exist."},
                    status=status.HTTP_404_NOT_FOUND
                )
            expense_type = ExpenseTypes.objects.get(id=pk)    
            serializer = ExpenseTypesSerializer(expense_type, data=request.data, partial=True)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            serializer.save()
            return Response({"message": f"Expense Types Updated successfully."}, status=status.HTTP_200_OK)      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while updating the Expense Types.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def delete(self, request, pk):
        try:
                        
            if not ExpenseTypes.objects.filter(id=pk).exists():
                return Response(
                    {"error": "Expense Types Does not Exist."},
                    status=status.HTTP_404_NOT_FOUND
                )
            ExpenseTypes.objects.get(id=pk).delete()
            if not ExpenseTypes.objects.filter(id=pk).exists():
                return Response({"message": f"Expense Types Deleted successfully."},
                    status=status.HTTP_204_NO_CONTENT   
                )
            else:
                return Response(
                    {"error": "Failed to delete an Expense Types."},
                    status=status.HTTP_400_BAD_REQUEST
                )      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Delete the Expense Types.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class OtherExpensesListCreateAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    # permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        try:
            
            other_expenses = OtherExpenses.objects.all().order_by('-id')
            serializer = OtherExpensesGetSerializer(other_expenses, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)              
                      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving Other Expenses.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def post(self, request, format=None):
        try:
          

            serializer = OtherExpensesSerializer(data=request.data)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            validated_data = serializer.validated_data
            # serializer.create(validated_data, user=request.user)
            serializer.create(validated_data)
            # return Response({"message": f"Other Expenses created successfully."}, status=status.HTTP_201_CREATED)
            return Response({"data": serializer.data}, status=status.HTTP_201_CREATED)
                      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while creating Other Expenses.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class OtherExpensesRetrieveUpdateDeleteAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    # permission_classes = (permissions.AllowAny,)
    def get(self, request, pk):
        try:
           
            if not OtherExpenses.objects.filter(id=pk).exists():
                return Response(
                    {"error": "Other Expenses Does not Exist."},
                    status=status.HTTP_404_NOT_FOUND
                )
            other_expenses = OtherExpenses.objects.get(id=pk)
            serializer = OtherExpensesGetSerializer(other_expenses)
            return Response(serializer.data, status=status.HTTP_200_OK)     
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving Other Expenses.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def put(self, request, pk):
        try:
                        
            if not OtherExpenses.objects.filter(id=pk).exists():
                return Response(
                    {"error": "Other Expenses Does not Exist."},
                    status=status.HTTP_404_NOT_FOUND
                )
            other_expenses = OtherExpenses.objects.get(id=pk)
            serializer = OtherExpensesSerializer(data=request.data)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            validated_data = serializer.validated_data
            serializer.update(other_expenses, validated_data)
            return Response({"message": f"Other Expenses Updated successfully."}, status=status.HTTP_200_OK)      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while updating Other Expenses  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
    def patch(self, request, pk):
        try:
                        
            if not OtherExpenses.objects.filter(id=pk).exists():
                return Response(
                    {"error": "Other Expenses Does not Exist."},
                    status=status.HTTP_404_NOT_FOUND
                )
            other_expenses = OtherExpenses.objects.get(id=pk)    
            serializer = OtherExpensesSerializer(other_expenses, data=request.data, partial=True)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            serializer.save()
            return Response({"message": f"Other Expenses Updated successfully."}, status=status.HTTP_200_OK)      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while updating Other Expenses.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def delete(self, request, pk):
        try:
                         
            if not OtherExpenses.objects.filter(id=pk).exists():
                return Response(
                    {"error": "Other Expenses Does not Exist."},
                    status=status.HTTP_404_NOT_FOUND
                )
            OtherExpenses.objects.get(id=pk).delete()
            if not OtherExpenses.objects.filter(id=pk).exists():
                return Response({"message": f"Other Expenses Deleted successfully."},
                    status=status.HTTP_204_NO_CONTENT   
                )
            else:
                return Response(
                    {"error": "Failed to delete Other Expenses."},
                    status=status.HTTP_400_BAD_REQUEST
                )      
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Delete Other Expenses.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class RetriveTotalProductCostAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request): 
        try:
           
     
            total_product_cost = Product.objects.aggregate(total_product_cost=Sum('buying_price'))        
            return Response(total_product_cost, status=status.HTTP_200_OK)         
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Total Product Cost.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ProductExcelReportAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request):
        try:
           
            report = Product.objects.all()
            serializer = ProductGetReportSerializer(report, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Product Report.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ProductsPerSupplierAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request, pk):
        try:
          
            """Retrieve all products belonging to a specific supplier."""
            products = Product.objects.filter(supplier_id=pk)
            if not products.exists():
                return Response({"message": "No products found for this supplier"}, status=status.HTTP_404_NOT_FOUND)
            
            serializer = ProductSerializer(products, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"An error occurred while retrieving products for this supplier: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )




class SalesPersonDashboardAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request):
        try:
            
            if not Order.objects.filter(user_email=user.email).exists():
                return Response(
                    {"error": "Order Does not Exist."},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            order = Order.objects.filter(user_email=user.email)
            serializer = OrderSerializer(order, many=True)
            # print(order)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while accessing Salesman Dashboard.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class RecentOrderLimitedAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request):
        try:
           
            recent_orders = Order.objects.all().order_by('-order_date')[:10]
            serializer = OrderSerializer(recent_orders, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while accessing Recent Orders.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

# ------------------------------------- Total Sales relative to Time --------------------------------------------------

class DailySalesAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request):
        try:
          
            today = timezone.now().date()
            orders = Order.objects.filter(order_date__date=today, status="Done", payment_status='Paid')
            total_sales = orders.aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
            serializer = OrderSerializer(orders, many=True)
            return Response({
                "date": str(today),
                "total_sales": total_sales,
                "orders": serializer.data
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"An error occurred while retrieving daily sales. {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
       
class WeeklySalesAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request):
        try:
            
            today = timezone.now().date()
            sales_data = []

            for i in range(6, -1, -1):  # Start from 6 days ago to today
                day = today - timedelta(days=i)
                orders = Order.objects.filter(
                    order_date__date=day,
                    status="Done", 
                    payment_status='Paid'
                )
                total_sales = orders.aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
                sales_data.append({
                    "period": day.strftime("%A"),  # Day name, e.g., "Monday"
                    "sales": float(total_sales)
                })

            return Response(sales_data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"An error occurred while retrieving last seven days sales. {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
class MonthlySalesAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request):
        try:
           
            today = timezone.now().date()
            year = today.year
            sales_data = []

            for month in range(1, 13):
                orders = Order.objects.filter(
                    order_date__year=year,
                    order_date__month=month,
                    status="Done", 
                    payment_status='Paid'
                )
                total_sales = orders.aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
                sales_data.append({
                    "period": calendar.month_name[month],
                    "sales": float(total_sales)
                })

            return Response(sales_data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"An error occurred while retrieving monthly sales. {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
class YearlySalesAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request):
        try:
           
            today = timezone.now().date()
            year = today.year

            orders = Order.objects.filter(
                order_date__year=year,
                status="Done", 
                payment_status='Paid'
            )
            total_sales = orders.aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0

            data = [{
                "period": str(year),
                "sales": float(total_sales)
            }]

            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"An error occurred while retrieving yearly sales. {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
# ------------------------------------- Total Sales relative to Time for Each User --------------------------------------------------

class DailySalesEachUserAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request):
        try:
           
            today = timezone.now().date()
            orders = Order.objects.filter(order_date__date=today, user_email=user.email, status="Done", payment_status='Paid')
            total_sales = orders.aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
            serializer = OrderSerializer(orders, many=True)
            return Response({
                "date": str(today),
                "total_sales": total_sales,
                "orders": serializer.data
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"An error occurred while retrieving daily sales. {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
       
class WeeklySalesEachUserAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request):
        try:
            
            today = timezone.now().date()
            sales_data = []

            for i in range(6, -1, -1):  # Start from 6 days ago to today
                day = today - timedelta(days=i)
                orders = Order.objects.filter(
                    order_date__date=day,
                    user_email=user.email,
                    status="Done", 
                    payment_status='Paid'
                )
                total_sales = orders.aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
                sales_data.append({
                    "period": day.strftime("%A"),  # Day name, e.g., "Monday"
                    "sales": float(total_sales)
                })

            return Response(sales_data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"An error occurred while retrieving last seven days sales. {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
class MonthlySalesEachUserAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request):
        try:
           
            today = timezone.now().date()
            year = today.year
            sales_data = []

            for month in range(1, 13):
                orders = Order.objects.filter(
                    order_date__year=year,
                    order_date__month=month,
                    user_email=user.email,
                    status="Done", 
                    payment_status='Paid'
                )
                total_sales = orders.aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
                sales_data.append({
                    "period": calendar.month_name[month],
                    "sales": float(total_sales)
                })

            return Response(sales_data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"An error occurred while retrieving monthly sales. {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
class YearlySalesEachUserAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request):
        try:
           
            today = timezone.now().date()
            year = today.year

            orders = Order.objects.filter(
                order_date__year=year,
                user_email=user.email,
                status="Done", 
                payment_status='Paid'
            )
            total_sales = orders.aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0

            data = [{
                "period": str(year),
                "sales": float(total_sales)
            }]

            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"An error occurred while retrieving yearly sales. {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ExportProductExcelAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request, *args, **kwargs):
        # Create workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"

        # Fetch data
        products = Product.objects.all().values(
            'id', 'name', 'description', 'package', 'piece', 'buying_price',
            'selling_price', 'unit', 'stock', 'receipt_no', 'user'
        )
        if not products:
            return Response({"error": "No product data available"}, status=204)

        # Write headers
        headers = list(products[0].keys())
        ws.append(headers)

        # Write data rows
        for product in products:
            ws.append(list(product.values()))

        # Prepare HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=products.xlsx'
        wb.save(response)
        return response

class ImportProductExcelAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            headers = [str(cell).strip() for cell in rows[0]]
            for row in rows[1:]:
                data = dict(zip(headers, row))
                # Adjust field names as needed for your Product model
                Product.objects.update_or_create(
                    id=data.get('id'),
                    defaults={
                        'name': data.get('name'),
                        'description': data.get('description'),
                        'package': data.get('package'),
                        'piece': data.get('piece'),
                        'buying_price': data.get('buying_price'),
                        'selling_price': data.get('selling_price'),
                        'unit': data.get('unit'),
                        'stock': data.get('stock'),
                        'receipt_no': data.get('receipt_no'),
                        'user': data.get('user'),  # Use FK id or handle lookup
                    }
                )
            return Response({"message": "Products imported successfully."}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": f"Failed to import products: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


class OrderLogListView(generics.ListAPIView):
    serializer_class = OrderPaymentLogSerializer

    def get_queryset(self):
        order_id = self.kwargs['order_id']
        return OrderPaymentLog.objects.filter(order_id=order_id).order_by('-timestamp')


class ProductLogAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]

    def get(self, request):
        try:
           
            log = ProductLog.objects.all()
            serializer = ProductLogSerializer(log, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Product Log.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



# class ProductWithBundleAPIView(APIView):
#     # permission_classes = [AllowAny]
#     def get(self, request): 
#         try:     
#             product_with_bundle = Product.objects.filter(is_bundle=True) 

#             serializer = ProductGetSerializer(product_with_bundle, many=True)
#             return Response(serializer.data, status=status.HTTP_200_OK)       
#         except KeyError as e:
#             return Response(
#                 {"error": f"An error occurred while Retriving the Product With Bundle.  {str(e)}"},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR
#             )


# class ProductWithOutBundleAPIView(APIView):
#     # permission_classes = [AllowAny]
#     def get(self, request): 
#         try:     
#             product_without_bundle = Product.objects.filter(is_bundle=False)

#             serializer = ProductGetSerializer(product_without_bundle, many=True)
#             return Response(serializer.data, status=status.HTTP_200_OK)       
#         except KeyError as e:
#             return Response(
#                 {"error": f"An error occurred while Retriving the Product With Out Bundle.  {str(e)}"},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR
#             )


class TotalOrderAPIView(APIView):
    # permission_classes = [AllowAny]
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    def get(self, request): 
        try:
           
            total_number_order = Order.objects.all().count()       
            return Response(total_number_order, status=status.HTTP_200_OK)

        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Total Order.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class TotalProductAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsTenantUser]
    # permission_classes = [AllowAny]
    def get(self, request): 
        try:     
            
            total_number = Product.objects.all().count()       
            return Response(total_number, status=status.HTTP_200_OK)         
        except KeyError as e:
            return Response(
                {"error": f"An error occurred while Retriving the Product.  {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )